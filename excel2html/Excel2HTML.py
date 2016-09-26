#################################################################################################x
# Excel2HTML provides a handy way of converting a provided CSV or Excel file into a NVD3.js 	#
# graph. The graph data itself is completely inline and links to publicly available JS 		#
# and CSS libraries. 										#
#												#
#												#
#												#
# Author: Shane Brennan										#
# Date: 20160926										#
# Version: 0.1											#
#################################################################################################


#!/usr/bin/env python
import csv
import sys
import os
import datetime
import re
import datetime
import openpyxl
from optparse import OptionParser

class Excel2HTML:

	def __init__(self, verbose, inputFilename):
		self.verbose = verbose
		
		if '.csv' in inputFilename[-4:].lower():
			self.processCSV(inputFilename)
		elif '.xls' in inputFilename[-4:].lower():
			self.processExcel(inputFilename)
		elif '.xlsx' in inputFilename[-5:].lower():
			self.processExcel(inputFilename)
		else:
			print 'Require valid Excel or CSV input File'
			exit(1)
	def getHTML(self):
		return 'test'

	def processCSV(self, inputFilename):
		inputFile = open(inputFilename, 'r')
		inputCSV = csv.reader(inputFile)
		index = 0
		for row in inputCSV:
			if index > 0:
				for element in row:
			else:
				header = row
			index += 1

	def processExcel(self, inputFilename):
		# Read in the Excel file as a list of lists
		workbook  = openpyxl.load_workbook(excelFilename, data_only=True)
		worksheet = workbook.worksheets[0]

		rowLimit = worksheet.max_row
		colLimit = worksheet.max_col
		for y in xrange(0, rowLimit):
			for x in xrange(1, colLimit:
				if worksheet.cell(row=x, column=y).value is not None:
					print worksheet.cell(row=x, column=1).value
		return 'test'

def main(argv):
	parser = OptionParser(usage="Usage: Excel2HTML <input-filename>")

        parser.add_option("-v", "--verbose",
                action="store_true",
                dest="verboseFlag",
                default=False,
                help="Verbose output from the script")

	(options, filename) = parser.parse_args()

	if len(filename) != 1 or not os.path.isfile(filename[0]) :
		print parser.print_help()
		exit(1)

	excel2HTML = Excel2HTML(options.verboseFlag, filename[0])
	print excel2HTML.getHTML()
		
if __name__ == "__main__":
    sys.exit(main(sys.argv))
