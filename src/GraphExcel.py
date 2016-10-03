#################################################################################################
# GraphExcel provides a handy way of converting a provided CSV or Excel file into a NVD3.js 	#
# graph. The graph data itself is completely inline and links to publicly available JS 		#
# and CSS libraries. 										#
#												#
#												#
#												#
# Author: Shane Brennan										#
# Date: 20160926										#
# Version: 0.1											#
#################################################################################################


#!/usr/bin/env python
import csv
import sys
import os
import datetime
import random
import re
import datetime
import logging
import openpyxl
from optparse import OptionParser

class GraphExcel:

	def __init__(self, inputFilename, verbose=False):
		self.logger   = logging.getLogger('graph_excel')
		self.rowSize  = 0
		self.colSize  = 0
		self.header   = []
		self.data     = []

		self.verbose = verbose
		if '.csv' in inputFilename[-4:].lower():
			self.logger.info('Processing CSV {0}'.format(inputFilename))
			self.processCSV(inputFilename)
		elif '.xls' in inputFilename[-4:].lower():
			self.logger.info('Processing Excel {0}'.format(inputFilename))
			self.processExcel(inputFilename)
		elif '.xlsx' in inputFilename[-5:].lower():
			self.logger.info('Processing Excel {0}'.format(inputFilename))
			self.processExcel(inputFilename)
		else:
			print 'Require valid Excel or CSV input File' 
			self.logger.error('Require valid Excel or CSV input File')

	def getHTML(self):
		return 'test'

	def processCSV(self, inputFilename):
		inputFile = open(inputFilename, 'r')
		inputCSV = csv.reader(inputFile)
		index = 0
		for row in inputCSV:
			if index == 0:
				self.header = row
				self.colSize = len(self.header)
			else:
				self.data.append(row)
			index += 1
		self.rowSize = index

	def processExcel(self, inputFilename):
		# Read in the Excel file as a list of lists
		workbook  = openpyxl.load_workbook(inputFilename, data_only=True)
		worksheet = workbook.worksheets[0]

		self.rowSize = worksheet.max_row
		self.colSize = worksheet.max_column
		for rowIndex in xrange(1, self.rowSize+1):
			row = []
			for colIndex in xrange(1, self.colSize+1):
				if worksheet.cell(row=rowIndex, column=colIndex).value is not None:
					element = worksheet.cell(row=rowIndex, column=colIndex).value
					row.append(element)
			if rowIndex == 1:
				self.header = row
			else:
				self.data.append(row)
	
	def getRows(self):
		return self.rowSize

	def getCols(self):
		return self.colSize

	def getHeader(self):
		return self.header

	def getData(self):
		return self.data

	def getRandomColours(self):
		coloursList = []
		for element in self.header:
			coloursList.append(self.getRandomHex())
		return coloursList

	def getRandomHex(self):
		digits = ['0','1','2','3','4','5','6','7','8','9','a','b','c','d','e','f']
		randomColour = ''
		for x in xrange(0,6):
			index = random.randint(0,15)
			randomColour += digits[index]
		return randomColour
def main(argv):
	parser = OptionParser(usage="Usage: Excel2HTML <input-filename>")

        parser.add_option("-v", "--verbose",
                action="store_true",
                dest="verboseFlag",
                default=False,
                help="Verbose output from the script")

	(options, filename) = parser.parse_args()

	if len(filename) != 1 or not os.path.isfile(filename[0]) :
		print parser.print_help()
		exit(1)

	excel2HTML = Excel2HTML(options.verboseFlag, filename[0])
	print excel2HTML.getHTML()
		
if __name__ == "__main__":
    sys.exit(main(sys.argv))
